"""
3자(N자) 비교 결과 → 엑셀 export.

시트 구성:
  - 비교매트릭스 : 패키지별 도구별 버전 나란히 + 존재도/판정 색상
  - 버전충돌     : 2개 이상 도구가 보유하지만 버전이 다른 것(가장 액션 필요)
  - {도구}만     : 각 도구에만 존재하는 패키지 (도구별 탭)
  - 요약         : 섹션별 개수 + 비율(%) + 텍스트 막대(█)로 한눈에

compare_multi() 결과를 받는다. 기존 2자 인터랙티브 엑셀(excel_interactive.py)과 별개.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from core.compare import DISPLAY_NAMES

import os
from dotenv import load_dotenv
load_dotenv()

HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12)
SECTION_FONT = Font(bold=True, color="1F497D")

# 판정별 행 색상
FILL_FULL_OK = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")   # 연녹: 전체일치
FILL_CONFLICT = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")   # 연빨: 버전충돌
FILL_PARTIAL = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")    # 연주황: 부분존재
FILL_UNIQUE = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")     # 회색: 단독

RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left")


def _disp(src: str) -> str:
    return DISPLAY_NAMES.get(src, src)


def _verdict(row: dict, n_sources: int) -> tuple[str, PatternFill]:
    cov, agr = row["coverage"], row["agreement"]
    if cov == n_sources and agr == "all_same":
        return "전체일치", FILL_FULL_OK
    if agr == "mismatch":
        return ("전체-버전충돌" if cov == n_sources else "부분-버전충돌"), FILL_CONFLICT
    if cov == 1:
        return "단독", FILL_UNIQUE
    return "부분(버전일치)", FILL_PARTIAL


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _set_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _bar(count: int, total: int, width: int = 24) -> str:
    """비율을 텍스트 막대(█/░)로. total=0이면 빈 막대."""
    frac = (count / total) if total else 0
    filled = round(frac * width)
    return "█" * filled + "░" * (width - filled)


def _pct(count: int, total: int) -> str:
    return f"{(100 * count / total):.1f}%" if total else "-"


def _versions_of(row: dict, src: str) -> str:
    if src not in row["present_in"]:
        return "-"
    return ", ".join(row["versions"][src]) or "(버전없음)"


def _paths_of(row: dict, src: str) -> str:
    """참고용 자산경로. 미보유 '-', 보유했지만 경로없음 '(경로없음)'."""
    fp = row.get("file_paths", {})
    if src not in fp:
        return "-"
    return "\n".join(fp[src]) if fp[src] else "(경로없음)"


WRAP = Alignment(wrap_text=True, vertical="top")


# ─────────────────────────── 개별 시트 ───────────────────────────
def _sheet_matrix(wb, result, sources):
    n = len(sources)
    ws = wb.create_sheet("비교매트릭스")
    # 버전 컬럼들 먼저 나란히, 그 다음 경로(자산) 컬럼들 나란히
    header = ["패키지명"]
    header += [f"{_disp(s)} 버전" for s in sources]
    first_path_col = len(header) + 1  # 경로 컬럼 시작 인덱스(1-based)
    header += [f"{_disp(s)} 경로" for s in sources]
    header += ["합계", "판정"]
    path_cols = list(range(first_path_col, first_path_col + n))
    ws.append(header)

    rows = sorted(
        result["rows"],
        key=lambda r: (0 if (r["coverage"] >= 2 and r["agreement"] == "mismatch") else 1,
                       -r["coverage"], r["name"]),
    )
    for r in rows:
        verdict, fill = _verdict(r, n)
        cells = [r["name"]]
        cells += [_versions_of(r, s) for s in sources]
        cells += [_paths_of(r, s) for s in sources]
        cells += [f"{r['coverage']}/{n}", verdict]
        ws.append(cells)
        for col in range(1, len(header) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill
        for col in path_cols:
            ws.cell(row=ws.max_row, column=col).alignment = WRAP

    _style_header(ws)
    ws.freeze_panes = "A2"
    widths = [40] + [20] * n + [34] * n + [8, 16]
    _set_widths(ws, widths)


def _sheet_conflicts(wb, result, sources):
    n = len(sources)
    ws = wb.create_sheet("버전충돌")
    header = ["패키지명", "합계"] + [f"{_disp(s)} 버전" for s in sources]
    ws.append(header)
    for r in result["version_conflicts"]:
        ws.append([r["name"], f"{r['coverage']}/{n}"] + [_versions_of(r, s) for s in sources])
        for col in range(1, len(header) + 1):
            ws.cell(row=ws.max_row, column=col).fill = FILL_CONFLICT
    _style_header(ws)
    ws.freeze_panes = "A2"
    _set_widths(ws, [42, 8] + [26] * n)


def _sheet_only_in(wb, result, src):
    """특정 도구에만 존재하는 패키지 시트."""
    ws = wb.create_sheet(f"{_disp(src)}만")
    ws.append(["패키지명", "버전", "경로"])
    for r in result["only_in"][src]:
        ws.append([r["name"], ", ".join(r["versions"][src]), _paths_of(r, src)])
        for c in range(1, 4):
            ws.cell(row=ws.max_row, column=c).fill = FILL_UNIQUE
        ws.cell(row=ws.max_row, column=3).alignment = WRAP
    _style_header(ws)
    ws.freeze_panes = "A2"
    _set_widths(ws, [46, 24, 40])


def _sheet_auto_merged(wb, auto_merged):
    """자동 병합 내역: 이름 표기만 다르고 버전이 완전히 같아 같은 패키지로 합친 것들."""
    ws = wb.create_sheet("자동병합")
    ws.append(["안내: 이름 표기만 다르고 버전이 완전히 같아 '같은 패키지'로 자동 병합한 내역 (검토 불필요).",
               "", "", "", ""])
    ws.cell(row=1, column=1).font = Font(italic=True, color="808080", size=9)
    ws.append(["원래 이름", "도구", "→ 병합된 대표 이름", "버전", "유사도"])
    for m in auto_merged:
        ws.append([m["from_name"], _disp(m["from_source"]), m["to_name"],
                   ", ".join(m["versions"]), m["score"]])
        for c in range(1, 6):
            ws.cell(row=ws.max_row, column=c).fill = FILL_FULL_OK
    _style_header(ws, row=2)
    ws.freeze_panes = "A3"
    _set_widths(ws, [34, 10, 34, 20, 8])


def _sheet_review(wb, candidates):
    """3자 퍼지 검토후보: 이름 표기만 다른 '같은 패키지 후보'. (자동 병합 X — 사람이 확인)

    왼쪽 = 한 도구에만 잡힌 '단독' 이름, 오른쪽 = 다른 도구(들)에 있는 매칭 후보.
    """
    ws = wb.create_sheet("검토후보")
    ws.append(["안내: 이름 표기가 달라 '단독'으로 잡혔지만 같은 패키지일 수 있는 후보. 유사도 높은 순.",
               "", "", "", "", "", "", ""])
    ws.cell(row=1, column=1).font = Font(italic=True, color="808080", size=9)
    ws.append(["단독 이름", "단독 도구", "버전", "↔ 매칭 후보", "후보 보유 도구", "버전", "유사도", "구분"])
    for c in candidates:
        ws.append([
            c["name_a"], _disp(c["source_a"]), ", ".join(c["versions_a"]),
            c["match_name"], " + ".join(_disp(s) for s in c["match_sources"]),
            ", ".join(c["match_versions"]),
            c["score"], c["band"],
        ])
    _style_header(ws, row=2)
    ws.freeze_panes = "A3"
    _set_widths(ws, [30, 12, 14, 30, 20, 14, 8, 20])


def _sheet_summary(wb, result, sources):
    n = len(sources)
    b = result["buckets"]
    total = len(result["rows"])
    in_all = len(b["full_match"]) + len(b["full_mismatch"])

    note_font = Font(italic=True, color="808080", size=9)
    sp_proj_name = os.getenv("PROJECT_KEY")
    bd_proj_name = os.getenv("BLACKDUCK_PROJECT_KEY")
    snyk_proj_name = os.getenv("SNYK_PROJECT_PATH")
    ws = wb.create_sheet("요약")
    ws.append(["3자 비교 요약", "", "", ""])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([f"분석 도구: {' / '.join(_disp(s) for s in sources)}", "", "", ""])
    ws.append([f"분석 프로젝트: {bd_proj_name} /  {sp_proj_name}  {snyk_proj_name}", "", "", ""])
    ws.append([f"분석 일자: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "", "", ""])
    
    
    # ws.cell(row=ws.max_row, column=1).font = note_font
    ws.append([])

    def section(title):
        ws.append([title, "", "", ""])
        r = ws.max_row
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = SECTION_FILL
        ws.cell(row=r, column=1).font = SECTION_FONT

    def note(text):
        ws.append([text, "", "", ""])
        ws.cell(row=ws.max_row, column=1).font = note_font

    def line(label, count, denom, indent=False):
        ws.append([("    " if indent else "") + label, count, _pct(count, denom), _bar(count, denom)])
        r = ws.max_row
        ws.cell(row=r, column=2).alignment = RIGHT
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=4).font = Font(name="Consolas")

    # 1) 커버리지 (몇 개 도구에 존재하나)
    section("① 커버리지")
    line(f"{n}개 도구 모두 존재", in_all, total)
    note("    → 세 도구가 모두 검출한 패키지")
    line("└ 버전 일치", len(b["full_match"]), total, indent=True)
    note("        → 검출도 됐고 버전 집합도 완전히 같음")
    line("└ 버전 충돌", len(b["full_mismatch"]), total, indent=True)
    note("        → 검출은 됐지만 도구마다 버전이 다름 → '버전충돌' 탭에서 확인 필요")
    if n > 2:
        line(f"일부 도구에만 존재", len(b["partial"]), total)
        note(f"    → {n}개 중 일부 도구만 검출 (한쪽이 놓쳤거나 이름 표기가 달라 매칭 실패 가능)")
    line("단독 존재", len(b["unique"]), total)
    note("    → 한 도구에서만 나온 패키지 → '○○만' 탭 참고 (그 도구만의 강점이거나 오탐일 수도 있음)")
    ws.append(["전체 패키지", total, "100.0%", ""])
    tr = ws.max_row
    ws.cell(row=tr, column=1).font = Font(bold=True)
    ws.cell(row=tr, column=2).font = Font(bold=True)
    ws.cell(row=tr, column=2).alignment = RIGHT
    note("    → 세 도구가 검출한 것을 합친 '고유 패키지' 총 개수 (같은 패키지는 1개로 셈)")
    ws.append([])

    # 2) 도구별 커버리지 (보유 / 단독)
    section("② 도구별 분포")
    note("총 검출 수 / 단독 검출 수")
    for s in sources:
        held = sum(1 for r in result["rows"] if s in r["present_in"])
        only = len(result["only_in"][s])
        ws.append([_disp(s), held, f"단독 {only}", _bar(held, total)])
        r = ws.max_row
        ws.cell(row=r, column=2).alignment = RIGHT
        ws.cell(row=r, column=4).font = Font(name="Consolas")
    ws.append([])

    # 3) 존재 조합별 분포
    section("③ 조합별 분포")
    for combo, cnt in sorted(result["combo_counts"].items(), key=lambda x: (-x[1], x[0])):
        label = " + ".join(_disp(s) for s in combo.split("+"))
        ws.append([label, cnt, _pct(cnt, total), _bar(cnt, total)])
        r = ws.max_row
        ws.cell(row=r, column=2).alignment = RIGHT
        ws.cell(row=r, column=3).alignment = RIGHT
        ws.cell(row=r, column=4).font = Font(name="Consolas")
    ws.append([])

    # 4) 용어 & 색상 (다른 탭에서 쓰는 판정/색 설명)
    section("④ 범례")
    for term, desc in [
        ("전체일치 (연녹)", "모든 도구에 존재 + 버전 집합 동일"),
        ("전체-버전충돌 (연빨)", "모든 도구에 존재하지만 버전이 다름"),
        ("부분-버전일치 (연주황)", "일부 도구에만 존재, 그 도구들끼리는 버전 동일"),
        ("부분-버전충돌 (연빨)", "일부 도구에만 존재 + 버전도 다름"),
        ("단독 (회색)", "한 도구에만 존재"),
    ]:
        ws.append([term, desc, "", ""])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=9)
        ws.cell(row=ws.max_row, column=2).font = Font(size=9)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 30


def export_matrix_excel(result: dict, candidates: list[dict] | None = None,
                        path: str = "comparison_multi.xlsx") -> str:
    sources = result["sources"]

    wb = Workbook()
    wb.remove(wb.active)

    _sheet_matrix(wb, result, sources)
    _sheet_conflicts(wb, result, sources)
    if result.get("auto_merged"):
        _sheet_auto_merged(wb, result["auto_merged"])
    if candidates:
        _sheet_review(wb, candidates)
    for s in sources:
        _sheet_only_in(wb, result, s)
    _sheet_summary(wb, result, sources)

    wb.active = 0  # 비교매트릭스를 기본 활성 시트로
    wb.save(path)
    print(f" 엑셀(3자) 저장 완료: {path}")
    return path
